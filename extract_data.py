"""
Extreu el llibre de la porra del Mundial a data.json per al web estàtic.

Ús:
    python extract_data.py "AA Classificació Porra Mundial 2026.xlsx"

Notes:
- L’script llegeix els resultats de fórmules desats a Excel/Google Sheets. Si els valors
  semblen desactualitzats, obre/recalcula el llibre abans o fes servir el mode en directe
  amb Google Sheets.
- No modifica el llibre.
"""
from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def as_int(value: Any, default: int = 0) -> int:
    value = clean(value)
    if value is None:
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def bounds(ws) -> tuple[int, int]:
    try:
        dim = ws.calculate_dimension(force=True)
    except TypeError:
        dim = ws.calculate_dimension()
    _, _, max_col, max_row = range_boundaries(dim)
    return max_row, max_col


def matrix(ws, max_rows: int | None = None, max_cols: int | None = None) -> list[list[Any]]:
    nrows, ncols = bounds(ws)
    if max_rows is not None:
        nrows = min(nrows, max_rows)
    if max_cols is not None:
        ncols = min(ncols, max_cols)
    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=1, max_row=nrows, min_col=1, max_col=ncols, values_only=True):
        rows.append([clean(v) for v in row])
    return rows


def cell(rows: list[list[Any]], r: int, c: int) -> Any:
    """1-based lookup from a matrix."""
    if r < 1 or c < 1 or r > len(rows) or c > len(rows[r - 1]):
        return None
    return rows[r - 1][c - 1]


def parse_leaderboard(rows: list[list[Any]]) -> list[dict[str, Any]]:
    headers = list(rows[0]) if rows else []
    if headers and headers[0] is None:
        headers[0] = "Participant"
    out: list[dict[str, Any]] = []
    for row in rows[1:250]:
        name = clean(row[0] if len(row) > 0 else None)
        if not name:
            continue
        entry = {
            "name": name,
            "paid": clean(row[1] if len(row) > 1 else None),
            "total": as_int(row[2] if len(row) > 2 else 0),
            "breakdown": {},
        }
        for idx, header in enumerate(headers[3:], start=3):
            if header is not None:
                entry["breakdown"][str(header)] = as_int(row[idx] if idx < len(row) else 0)
        out.append(entry)
    out.sort(key=lambda x: (-x["total"], x["name"].lower()))
    last_total = None
    rank = 0
    for idx, row in enumerate(out, start=1):
        if row["total"] != last_total:
            rank = idx
            last_total = row["total"]
        row["rank"] = rank
    return out


def parse_stats(rows: list[list[Any]]) -> dict[str, Any]:
    predictions: list[dict[str, Any]] = []
    for row in rows[1:250]:
        name = clean(row[0] if len(row) > 0 else None)
        if not name:
            continue
        predictions.append({
            "name": name,
            "champion": clean(row[1] if len(row) > 1 else None),
            "runnerUp": clean(row[2] if len(row) > 2 else None),
            "third": clean(row[3] if len(row) > 3 else None),
            "fourth": clean(row[4] if len(row) > 4 else None),
            "topScorer": clean(row[5] if len(row) > 5 else None),
            "topScorerGoals": clean(row[6] if len(row) > 6 else None),
        })

    team_summary: list[dict[str, Any]] = []
    for row in rows[1:250]:
        team = clean(row[8] if len(row) > 8 else None)
        if not team:
            continue
        team_summary.append({
            "team": team,
            "champion": as_int(row[9] if len(row) > 9 else 0),
            "runnerUp": as_int(row[10] if len(row) > 10 else 0),
            "third": as_int(row[11] if len(row) > 11 else 0),
            "fourth": as_int(row[12] if len(row) > 12 else 0),
            "total": as_int(row[13] if len(row) > 13 else 0),
        })
    team_summary.sort(key=lambda x: (-x["total"], x["team"]))

    scorer_summary: list[dict[str, Any]] = []
    for row in rows[1:250]:
        scorer = clean(row[14] if len(row) > 14 else None)
        if not scorer:
            continue
        scorer_summary.append({"player": scorer, "count": as_int(row[15] if len(row) > 15 else 0)})
    scorer_summary.sort(key=lambda x: (-x["count"], x["player"]))

    return {"predictions": predictions, "teamSummary": team_summary, "scorerSummary": scorer_summary}


def parse_groups(rows: list[list[Any]]) -> list[dict[str, Any]]:
    groups: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=1):
        label = clean(row[0] if row else None)
        if not isinstance(label, str) or not label.upper().startswith("GROUP "):
            continue
        group = {"name": label, "matches": [], "standings": []}
        for mr in range(idx + 3, idx + 9):
            home = cell(rows, mr, 1)
            away = cell(rows, mr, 6)
            if home and away:
                group["matches"].append({
                    "home": home,
                    "homeScore": cell(rows, mr, 3),
                    "awayScore": cell(rows, mr, 5),
                    "away": away,
                    "winner": cell(rows, mr, 8),
                })
        for sr in range(idx + 12, idx + 16):
            team = cell(rows, sr, 2)
            if team:
                group["standings"].append({
                    "pos": as_int(cell(rows, sr, 1)),
                    "team": team,
                    "played": as_int(cell(rows, sr, 3)),
                    "wins": as_int(cell(rows, sr, 4)),
                    "draws": as_int(cell(rows, sr, 5)),
                    "losses": as_int(cell(rows, sr, 6)),
                    "gf": as_int(cell(rows, sr, 7)),
                    "ga": as_int(cell(rows, sr, 8)),
                    "gd": as_int(cell(rows, sr, 9)),
                    "pts": as_int(cell(rows, sr, 10)),
                    "rank": as_int(cell(rows, sr, 11)),
                })
        groups.append(group)
    return groups


def parse_thirds(rows: list[list[Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for r in range(5, min(len(rows), 80) + 1):
        group = cell(rows, r, 1)
        team = cell(rows, r, 2)
        if not group or not team:
            continue
        if not (isinstance(group, str) and group.startswith("3")):
            continue
        out.append({
            "group": group,
            "team": team,
            "played": as_int(cell(rows, r, 3)),
            "wins": as_int(cell(rows, r, 4)),
            "draws": as_int(cell(rows, r, 5)),
            "losses": as_int(cell(rows, r, 6)),
            "gf": as_int(cell(rows, r, 7)),
            "ga": as_int(cell(rows, r, 8)),
            "gd": as_int(cell(rows, r, 9)),
            "pts": as_int(cell(rows, r, 10)),
            "qualified": cell(rows, r, 11),
            "rank": as_int(cell(rows, r, 12)),
        })
    out.sort(key=lambda x: (x["rank"], -x["pts"], -x["gd"], x["team"]))
    return out


def parse_knockout(rows: list[list[Any]]) -> dict[str, Any]:
    round_names = {
        "ROUND OF 32": "Setzens de final",
        "ROUND OF 16": "Vuitens de final",
        "QUARTER-FINALS": "Quarts de final",
        "SEMI-FINALS": "Semifinals",
        "THIRD PLACE MATCH": "Partit pel tercer lloc",
        "🏆 FINAL 🏆": "Final",
    }
    rounds: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for idx, row in enumerate(rows[:140], start=1):
        label = clean(row[0] if row else None)
        if isinstance(label, str) and label in round_names:
            current = {"name": round_names[label], "matches": []}
            rounds.append(current)
            continue
        if current is None or not isinstance(label, str) or re.fullmatch(r"M\d+", label) is None:
            continue
        current["matches"].append({
            "id": label,
            "home": cell(rows, idx, 2),
            "homeSeed": cell(rows, idx, 3),
            "homeScore": cell(rows, idx, 4),
            "awayScore": cell(rows, idx, 5),
            "away": cell(rows, idx, 6),
            "awaySeed": cell(rows, idx, 7),
            "winner": cell(rows, idx, 8),
            "penHome": cell(rows, idx, 9),
            "penAway": cell(rows, idx, 10),
            "otherResult": cell(rows, idx, 11),
        })

    qualifiers = []
    for r in range(6, 18):
        group = cell(rows, r, 1)
        if group:
            qualifiers.append({
                "group": group,
                "winner": cell(rows, r, 2),
                "runnerUp": cell(rows, r, 4),
                "third": cell(rows, r, 6),
                "thirdPts": as_int(cell(rows, r, 7)),
                "thirdGd": as_int(cell(rows, r, 8)),
            })
    return {
        "rounds": rounds,
        "results": {
            "fourth": cell(rows, 77, 2),
            "third": cell(rows, 78, 2),
            "runnerUp": cell(rows, 79, 2),
            "champion": cell(rows, 80, 2),
            "topScorer": cell(rows, 79, 6),
            "topScorerGoals": cell(rows, 80, 6),
        },
        "qualifiers": qualifiers,
    }


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        raise SystemExit(2)
    workbook_path = Path(sys.argv[1]).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    data = {
        "meta": {
            "sourceFile": workbook_path.name,
            "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "mode": "xlsx-cache",
        },
        "leaderboard": parse_leaderboard(matrix(wb["Classificació"], max_cols=30)),
        "stats": parse_stats(matrix(wb["Estadística"], max_cols=20)),
        "groups": parse_groups(matrix(wb["Fase Grups"], max_cols=12)),
        "thirdPlaces": parse_thirds(matrix(wb["3ers Classificats"], max_cols=12)),
        "knockout": parse_knockout(matrix(wb["Eliminatòries"], max_cols=12)),
    }
    out = Path(__file__).with_name("data.json")
    out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {out}")
    print(f"Participants: {len(data['leaderboard'])}; groups: {len(data['groups'])}; rounds: {len(data['knockout']['rounds'])}")


if __name__ == "__main__":
    main()
